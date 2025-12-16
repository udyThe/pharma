"""
Report Generator Service
Generate PDF and Excel reports from agent analysis.
"""
import os
import re
from datetime import datetime
from pathlib import Path
from typing import Dict, List, Optional, Any
import json

try:
    from fpdf import FPDF
    FPDF_AVAILABLE = True
except ImportError:
    FPDF_AVAILABLE = False
    print("fpdf2 not installed. PDF generation disabled.")

try:
    import openpyxl
    from openpyxl.styles import Font, Alignment, PatternFill, Border, Side
    from openpyxl.utils import get_column_letter
    OPENPYXL_AVAILABLE = True
except ImportError:
    OPENPYXL_AVAILABLE = False
    print("openpyxl not installed. Excel generation disabled.")


def clean_text_for_pdf(text: str) -> str:
    """Clean text for PDF generation - handle unicode and special characters."""
    if not text:
        return ""
    # Replace common unicode characters with ASCII equivalents
    replacements = {
        '\u2022': '-',  # bullet
        '\u2019': "'",  # right single quote
        '\u2018': "'",  # left single quote
        '\u201c': '"',  # left double quote
        '\u201d': '"',  # right double quote
        '\u2013': '-',  # en dash
        '\u2014': '-',  # em dash
        '\u2026': '...',  # ellipsis
        '\u00a0': ' ',  # non-breaking space
        '\u00b7': '-',  # middle dot
        '\u2192': '->',  # right arrow
        '\u2190': '<-',  # left arrow
        '\u2713': '[x]',  # checkmark
        '\u2717': '[ ]',  # cross mark
        '\u00ae': '(R)',  # registered
        '\u2122': '(TM)',  # trademark
        '\u00a9': '(C)',  # copyright
        '\u00b0': ' degrees',  # degree symbol
        '\u00b1': '+/-',  # plus-minus
        '\u00d7': 'x',  # multiplication
        '\u00f7': '/',  # division
        '\u221e': 'infinity',  # infinity
        '\u2264': '<=',  # less than or equal
        '\u2265': '>=',  # greater than or equal
        '•': '-',
        '→': '->',
        '←': '<-',
        '✓': '[x]',
        '✗': '[ ]',
        '★': '*',
        '☆': '*',
    }
    for unicode_char, ascii_char in replacements.items():
        text = text.replace(unicode_char, ascii_char)
    
    # Remove any remaining non-ASCII characters or replace with ?
    text = text.encode('ascii', 'replace').decode('ascii')
    return text


class ReportGenerator:
    """Generate professional reports from agent analysis."""
    
    def __init__(self, output_dir: Optional[str] = None):
        """Initialize the report generator."""
        self.base_dir = Path(__file__).resolve().parent.parent.parent
        self.output_dir = Path(output_dir) if output_dir else self.base_dir / "reports"
        self.output_dir.mkdir(parents=True, exist_ok=True)
    
    def generate_pdf(
        self,
        title: str,
        query: str,
        content: str,
        metadata: Optional[Dict] = None
    ) -> str:
        """
        Generate a PDF report.
        
        Args:
            title: Report title
            query: Original user query
            content: Main report content (markdown-like text)
            metadata: Optional metadata (date, agents used, etc.)
        
        Returns:
            Path to generated PDF file
        """
        if not FPDF_AVAILABLE:
            return "PDF generation not available. Install fpdf2: pip install fpdf2"
        
        try:
            pdf = FPDF()
            pdf.set_auto_page_break(auto=True, margin=15)
            pdf.add_page()
            
            # Clean text for PDF
            title = clean_text_for_pdf(title)
            query = clean_text_for_pdf(query)
            content = clean_text_for_pdf(content)
            
            # Title
            pdf.set_font("Helvetica", "B", 20)
            pdf.set_text_color(0, 51, 102)  # Dark blue
            pdf.cell(0, 15, title, ln=True, align="C")
            pdf.ln(5)
            
            # Metadata bar
            pdf.set_font("Helvetica", "", 10)
            pdf.set_text_color(100, 100, 100)
            timestamp = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            pdf.cell(0, 8, f"Generated: {timestamp}", ln=True)
            
            if metadata:
                if metadata.get("agents_used"):
                    agents = [clean_text_for_pdf(str(a)) for a in metadata['agents_used']]
                    pdf.cell(0, 8, f"Agents Used: {', '.join(agents)}", ln=True)
                if metadata.get("user"):
                    pdf.cell(0, 8, f"User: {clean_text_for_pdf(str(metadata['user']))}", ln=True)
            
            pdf.ln(5)
            
            # Query section
            pdf.set_fill_color(240, 240, 240)
            pdf.set_font("Helvetica", "B", 12)
            pdf.set_text_color(0, 0, 0)
            pdf.cell(0, 10, "Original Query:", ln=True)
            pdf.set_font("Helvetica", "I", 11)
            pdf.multi_cell(0, 7, query)
            pdf.ln(5)
            
            # Separator line
            pdf.set_draw_color(0, 51, 102)
            pdf.line(10, pdf.get_y(), 200, pdf.get_y())
            pdf.ln(10)
            
            # Section: Analysis Results
            pdf.set_font("Helvetica", "B", 14)
            pdf.set_text_color(0, 51, 102)
            pdf.cell(0, 10, "Analysis Results", ln=True)
            pdf.ln(3)
            
            # Main content
            pdf.set_font("Helvetica", "", 11)
            pdf.set_text_color(0, 0, 0)
            
            # Process content line by line
            lines = content.split("\n")
            for line in lines:
                line = line.strip()
                
                if not line:
                    pdf.ln(3)
                    continue
                
                # Handle markdown headers (##, ###)
                if line.startswith("### "):
                    pdf.set_font("Helvetica", "B", 11)
                    pdf.set_text_color(0, 51, 102)
                    pdf.multi_cell(0, 7, line[4:].replace("**", ""))
                    pdf.set_font("Helvetica", "", 11)
                    pdf.set_text_color(0, 0, 0)
                elif line.startswith("## "):
                    pdf.set_font("Helvetica", "B", 12)
                    pdf.set_text_color(0, 51, 102)
                    pdf.multi_cell(0, 8, line[3:].replace("**", ""))
                    pdf.set_font("Helvetica", "", 11)
                    pdf.set_text_color(0, 0, 0)
                # Handle bold headers
                elif line.startswith("**") and line.endswith("**"):
                    pdf.set_font("Helvetica", "B", 12)
                    pdf.set_text_color(0, 51, 102)
                    pdf.multi_cell(0, 8, line.replace("**", ""))
                    pdf.set_font("Helvetica", "", 11)
                    pdf.set_text_color(0, 0, 0)
                
                # Handle bullet points
                elif line.startswith("- ") or line.startswith("* "):
                    pdf.set_x(15)
                    text = line[2:].replace("**", "").replace("*", "")
                    pdf.multi_cell(0, 6, f"- {text}")
                
                # Handle numbered items
                elif len(line) > 2 and line[0].isdigit() and line[1:3] in [". ", ") "]:
                    pdf.set_x(15)
                    pdf.multi_cell(0, 6, line.replace("**", "").replace("*", ""))
                
                # Regular text
                else:
                    clean_line = line.replace("**", "").replace("*", "")
                    pdf.multi_cell(0, 6, clean_line)
            
            # Footer
            pdf.ln(10)
            pdf.set_font("Helvetica", "I", 9)
            pdf.set_text_color(150, 150, 150)
            pdf.cell(0, 10, "Generated by Pharma Agentic AI System", ln=True, align="C")
            
            # Save PDF
            filename = f"pharma_report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.pdf"
            filepath = self.output_dir / filename
            pdf.output(str(filepath))
            
            return str(filepath)
        
        except Exception as e:
            import traceback
            return f"Error generating PDF: {str(e)}\n{traceback.format_exc()}"
    
    def generate_excel(
        self,
        title: str,
        query: str,
        data: Dict[str, Any],
        metadata: Optional[Dict] = None
    ) -> str:
        """
        Generate an Excel report with structured data.
        
        Args:
            title: Report title
            query: Original user query
            data: Dictionary containing structured data for different sections
            metadata: Optional metadata
        
        Returns:
            Path to generated Excel file
        """
        if not OPENPYXL_AVAILABLE:
            return "Excel generation not available. Install openpyxl: pip install openpyxl"
        
        try:
            wb = openpyxl.Workbook()
            
            # Styles
            header_font = Font(bold=True, size=14, color="FFFFFF")
            header_fill = PatternFill(start_color="003366", end_color="003366", fill_type="solid")
            subheader_font = Font(bold=True, size=11)
            subheader_fill = PatternFill(start_color="E0E0E0", end_color="E0E0E0", fill_type="solid")
            thin_border = Border(
                left=Side(style='thin'),
                right=Side(style='thin'),
                top=Side(style='thin'),
                bottom=Side(style='thin')
            )
            
            # Summary Sheet
            ws = wb.active
            ws.title = "Summary"
            
            # Title
            ws.merge_cells("A1:F1")
            ws["A1"] = title
            ws["A1"].font = Font(bold=True, size=18, color="003366")
            ws["A1"].alignment = Alignment(horizontal="center")
            
            # Metadata
            ws["A3"] = "Generated:"
            ws["B3"] = datetime.now().strftime("%Y-%m-%d %H:%M:%S")
            ws["A4"] = "Query:"
            ws["B4"] = query
            ws["B4"].alignment = Alignment(wrap_text=True)
            ws.row_dimensions[4].height = 40
            
            if metadata and metadata.get("agents_used"):
                ws["A5"] = "Agents Used:"
                ws["B5"] = ", ".join(metadata["agents_used"])
            
            # Main findings
            row = 7
            ws[f"A{row}"] = "Key Findings"
            ws[f"A{row}"].font = header_font
            ws[f"A{row}"].fill = header_fill
            ws.merge_cells(f"A{row}:F{row}")
            
            row += 1
            if "findings" in data:
                for finding in data["findings"]:
                    ws[f"A{row}"] = f"• {finding}"
                    ws[f"A{row}"].alignment = Alignment(wrap_text=True)
                    ws.merge_cells(f"A{row}:F{row}")
                    row += 1
            
            # Recommendations
            row += 1
            ws[f"A{row}"] = "Recommendations"
            ws[f"A{row}"].font = header_font
            ws[f"A{row}"].fill = header_fill
            ws.merge_cells(f"A{row}:F{row}")
            
            row += 1
            if "recommendations" in data:
                for rec in data["recommendations"]:
                    ws[f"A{row}"] = f"• {rec}"
                    ws[f"A{row}"].alignment = Alignment(wrap_text=True)
                    ws.merge_cells(f"A{row}:F{row}")
                    row += 1
            
            # Set column widths
            ws.column_dimensions["A"].width = 20
            ws.column_dimensions["B"].width = 50
            
            # Data Sheets
            if "market_data" in data:
                self._add_data_sheet(wb, "Market Data", data["market_data"], header_fill, header_font, thin_border)
            
            if "patent_data" in data:
                self._add_data_sheet(wb, "Patent Data", data["patent_data"], header_fill, header_font, thin_border)
            
            if "trial_data" in data:
                self._add_data_sheet(wb, "Clinical Trials", data["trial_data"], header_fill, header_font, thin_border)
            
            if "competitor_data" in data:
                self._add_data_sheet(wb, "Competitor Intel", data["competitor_data"], header_fill, header_font, thin_border)
            
            # Save Excel
            filename = f"report_{datetime.now().strftime('%Y%m%d_%H%M%S')}.xlsx"
            filepath = self.output_dir / filename
            wb.save(str(filepath))
            
            return str(filepath)
        
        except Exception as e:
            return f"Error generating Excel: {str(e)}"
    
    def _add_data_sheet(self, wb, sheet_name: str, data: List[Dict], header_fill, header_font, border):
        """Add a data sheet to the workbook."""
        ws = wb.create_sheet(title=sheet_name)
        
        if not data:
            ws["A1"] = "No data available"
            return
        
        # Get all keys from data
        all_keys = set()
        for row in data:
            all_keys.update(row.keys())
        headers = sorted(list(all_keys))
        
        # Write headers
        for col, header in enumerate(headers, 1):
            cell = ws.cell(row=1, column=col, value=header.replace("_", " ").title())
            cell.font = header_font
            cell.fill = header_fill
            cell.border = border
            ws.column_dimensions[get_column_letter(col)].width = max(15, len(header) + 5)
        
        # Write data
        for row_idx, row_data in enumerate(data, 2):
            for col_idx, header in enumerate(headers, 1):
                value = row_data.get(header, "")
                if isinstance(value, list):
                    value = ", ".join(str(v) for v in value)
                cell = ws.cell(row=row_idx, column=col_idx, value=value)
                cell.border = border
                cell.alignment = Alignment(wrap_text=True)


def generate_pdf_report(title: str, query: str, content: str, metadata: Optional[Dict] = None) -> str:
    """Convenience function to generate PDF report."""
    generator = ReportGenerator()
    return generator.generate_pdf(title, query, content, metadata)


def generate_excel_report(title: str, query: str, data: Dict[str, Any], metadata: Optional[Dict] = None) -> str:
    """Convenience function to generate Excel report."""
    generator = ReportGenerator()
    return generator.generate_excel(title, query, data, metadata)
